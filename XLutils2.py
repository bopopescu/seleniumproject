import openpyxl

def rowcount(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.max_row

def colmncount(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.max_column

def readdata(path,sheetname,rowcount,columncount):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rowcount,column=columncount).value

def writedata(path,sheetname,rowcount,columncount,data):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rowcount,column=columncount).value=data
    workbook.save(path)
