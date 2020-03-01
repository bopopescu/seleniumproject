import openpyxl
path="C:/Users/amruth/Desktop/Hemraj/Login1.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
rows=sheet.max_row
cols=sheet.max_column

for r in range(1,rows+1):
    for c in range(1,cols+1):
        data=sheet.cell(row=r,column=c).value
        print(data,end='    ')

    print()
