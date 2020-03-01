import openpyxl

path="C:/Users/amruth/Desktop/Hemraj/Read.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
column=sheet.max_column
row=sheet.max_row
print(row)
print(column)

for r in range(1,row+1):
    for c in range(1,column+1):
        print(sheet.cell(row=r,column=c).value,end='       ')
    print()

